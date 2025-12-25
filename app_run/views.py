from django.conf import settings
from django.contrib.auth.models import User
from django.db.models import Min, Max, Q, Count, Avg, Sum
from django.shortcuts import get_object_or_404

from geopy.distance import geodesic
from django_filters.rest_framework import DjangoFilterBackend
from openpyxl import load_workbook

from rest_framework import viewsets, generics
from rest_framework import serializers
from rest_framework.decorators import api_view, action, permission_classes
from rest_framework.filters import SearchFilter, OrderingFilter
from rest_framework.permissions import AllowAny
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.viewsets import ReadOnlyModelViewSet

from .models import (
    Run,
    AthleteInfo,
    Challenge,
    Position,
    CollectibleItem,
    Subscribe,
)
from .serializers import (
    RunSerializer,
    AthleteInfoSerializer,
    ChallengeSerializer,
    PositionSerializer,
    CollectibleItemSerializer,
    UserBaseSerializer,
    AthleteDetailSerializer,
    CoachDetailSerializer,
    RateCoachSerializer,
)
from .pagination import CustomPageNumberPagination


"""API представления и ViewSet-ы бегового трекера."""


@api_view(["GET"])
def company_details(request):
    data = {
        "company_name": getattr(settings, "COMPANY_NAME", "Company"),
        "slogan": getattr(settings, "COMPANY_SLOGAN", ""),
        "contacts": getattr(settings, "COMPANY_CONTACTS", ""),
    }
    return Response(data)


@api_view(["POST"])
@permission_classes([AllowAny])
def subscribe_to_coach(request, id):
    """
    Оформляет подписку атлета на тренера.
    POST /api/subscribe_to_coach/<coach_id>/
    body: {"athlete": <athlete_id>}
    """

    # 1. Тренер существует?
    coach = get_object_or_404(User, pk=id)

    # Тренером считаем is_staff == True
    if not coach.is_staff:
        return Response({"error": "User is not a coach"}, status=400)

    # 2. Достаём id атлета из тела
    athlete_id = request.data.get("athlete")
    if athlete_id is None:
        return Response({"error": "Field 'athlete' is required"}, status=400)

    try:
        athlete = User.objects.get(pk=athlete_id)
    except User.DoesNotExist:
        return Response({"error": "Athlete not found"}, status=400)

    # Атлет — тот, кто не is_staff
    if athlete.is_staff:
        return Response({"error": "User is not an athlete"}, status=400)

    # 3. Уже подписан?
    if Subscribe.objects.filter(athlete=athlete, coach=coach).exists():
        return Response({"error": "Already subscribed"}, status=400)

    # 4. Создаём подписку
    Subscribe.objects.create(athlete=athlete, coach=coach)
    return Response({"status": "ok"}, status=200)


@api_view(["GET"])
def challenges_summary(request):
    """
    /api/challenges_summary/
    Возвращает список челленджей с атлетами, которые их выполнили.
    """
    # 1 запрос в БД, подтягиваем сразу связанного атлета
    challenges = Challenge.objects.select_related("athlete").order_by(
        "full_name", "athlete__id"
    )

    # Группируем по названию челленджа
    summary = {}

    for ch in challenges:
        name = ch.full_name
        athlete = ch.athlete

        if name not in summary:
            summary[name] = []

        summary[name].append(
            {
                "id": athlete.id,
                "full_name": f"{athlete.first_name} {athlete.last_name}",
                "username": athlete.username,
            }
        )

    # Преобразуем в список, как ждёт фронт
    result = [
        {
            "name_to_display": name,
            "athletes": athletes,
        }
        for name, athletes in summary.items()
    ]

    return Response(result)


@api_view(["POST"])
@permission_classes([AllowAny])
def rate_coach(request, coach_id):
    """Устанавливает или обновляет оценку тренера атлетом."""

    # 1. Проверяем тренера
    coach = get_object_or_404(User, pk=coach_id)
    if not coach.is_staff:
        return Response({"error": "User is not a coach"}, status=400)

    serializer = RateCoachSerializer(data=request.data)
    serializer.is_valid(raise_exception=True)

    athlete_id = serializer.validated_data["athlete"]
    rating = serializer.validated_data["rating"]

    # 2. Проверяем атлета
    try:
        athlete = User.objects.get(pk=athlete_id)
    except User.DoesNotExist:
        return Response({"error": "Athlete not found"}, status=400)

    if athlete.is_staff:
        return Response({"error": "User is not an athlete"}, status=400)

    # 3. Проверяем подписку
    try:
        sub = Subscribe.objects.get(athlete=athlete, coach=coach)
    except Subscribe.DoesNotExist:
        return Response(
            {"error": "Athlete is not subscribed to this coach"}, status=400
        )

    # 4. Ставим или обновляем рейтинг
    sub.rating = rating
    sub.save(update_fields=["rating"])

    return Response({"status": "ok", "rating": rating})


@api_view(["GET"])
@permission_classes([AllowAny])
def analytics_for_coach(request, coach_id):
    """Возвращает аналитические показатели атлетов тренера."""

    get_object_or_404(User, pk=coach_id)

    # список атлетов, подписанных на тренера
    athlete_ids = Subscribe.objects.filter(coach_id=coach_id).values_list(
        "athlete_id", flat=True
    )

    # завершённые забеги этих атлетов
    runs = Run.objects.filter(
        athlete_id__in=athlete_ids,
        status=Run.Status.FINISHED,
    )

    # --- longest_run ---
    longest = runs.values("athlete_id", "distance").order_by("-distance").first()

    # --- total_run ---
    total = (
        runs.values("athlete_id")
        .annotate(total_distance=Sum("distance"))
        .order_by("-total_distance")
        .first()
    )

    # --- speed_avg ---
    speed = (
        runs.filter(speed__isnull=False)
        .values("athlete_id")
        .annotate(avg_speed=Avg("speed"))
        .order_by("-avg_speed")
        .first()
    )

    data = {
        "longest_run_user": longest["athlete_id"] if longest else None,
        "longest_run_value": float(longest["distance"]) if longest else None,
        "total_run_user": total["athlete_id"] if total else None,
        "total_run_value": float(total["total_distance"]) if total else None,
        "speed_avg_user": speed["athlete_id"] if speed else None,
        "speed_avg_value": float(speed["avg_speed"]) if speed else None,
    }

    return Response(data)


class RunViewSet(viewsets.ModelViewSet):
    """API для управления забегами атлетов."""

    serializer_class = RunSerializer
    filter_backends = [DjangoFilterBackend, OrderingFilter]
    filterset_fields = ["status", "athlete"]
    ordering_fields = ["created_at"]
    pagination_class = None

    def get_queryset(self):
        qs = Run.objects.select_related("athlete")
        athlete_id = self.request.query_params.get("athlete_id")
        if athlete_id:
            qs = qs.filter(athlete_id=athlete_id)
        return qs

    @action(detail=True, methods=["post"])
    def start(self, request, pk=None):
        """
        Запускает забег и переводит его в статус IN_PROGRESS.
        POST /api/runs/<id>/start/
        """
        run = self.get_object()

        if run.status != Run.Status.INIT:
            return Response(
                {"error": "Забег уже запущен или завершён"},
                status=400,
            )

        run.status = Run.Status.IN_PROGRESS
        run.save()
        return Response({"status": run.status})

    @action(detail=True, methods=["post"])
    def stop(self, request, pk=None):
        """
        Завершает забег и рассчитывает итоговые показатели.
        POST /api/runs/<id>/stop/
        """
        run = self.get_object()

        if run.status != Run.Status.IN_PROGRESS:
            return Response(
                {"error": "Забег ещё не запущен или уже завершён"},
                status=400,
            )

        run.status = Run.Status.FINISHED
        run.save()

        # После завершения — считаем время и среднюю скорость
        self.calculate_run_time(run)
        return Response({"status": run.status})

    def list(self, request, *args, **kwargs):
        """
        Если есть ?size=... → включаем пагинацию.
        Иначе отдаём весь список.
        """
        queryset = self.filter_queryset(self.get_queryset())

        if "size" in request.query_params:
            paginator = CustomPageNumberPagination()
            page = paginator.paginate_queryset(queryset, request, view=self)
            serializer = self.get_serializer(page, many=True)
            return paginator.get_paginated_response(serializer.data)

        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)

    def calculate_run_time(self, run: Run):
        """
        Время забега = max(date_time) - min(date_time) по позициям.
        Заодно считаем среднюю скорость по Position.speed.
        """
        agg = run.positions.aggregate(
            min_dt=Min("date_time"),
            max_dt=Max("date_time"),
        )
        min_dt = agg["min_dt"]
        max_dt = agg["max_dt"]

        if not (min_dt and max_dt):
            return

        run.run_time_seconds = int((max_dt - min_dt).total_seconds())

        positions = run.positions.exclude(speed__isnull=True)
        if positions.exists():
            avg_speed = round(
                sum(p.speed for p in positions) / positions.count(),
                2,
            )
            run.speed = avg_speed

        # Вызов save() триггерит нашу бизнес-логику в models.Run.save()
        run.save(update_fields=["run_time_seconds", "speed"])


class UserViewSet(ReadOnlyModelViewSet):
    """API для просмотра пользователей приложения."""

    filter_backends = [SearchFilter, OrderingFilter]
    search_fields = ["first_name", "last_name"]
    ordering_fields = ["date_joined"]
    pagination_class = None

    def get_queryset(self):
        qs = User.objects.exclude(is_superuser=True)

        user_type = self.request.query_params.get("type")
        if user_type == "coach":
            qs = qs.filter(is_staff=True)
        elif user_type == "athlete":
            qs = qs.filter(is_staff=False)

        return qs.annotate(
            runs_finished=Count("run", filter=Q(run__status=Run.Status.FINISHED)),
            rating=Avg("subscribers__rating"),
        )

    def get_serializer_class(self):
        # /api/users/  → базовый сериализатор без coach/athletes
        if self.action == "list":
            return UserBaseSerializer

        # /api/users/<id>/ → разные сериализаторы для тренера и атлета
        if self.action == "retrieve":
            user = self.get_object()
            if user.is_staff:
                return CoachDetailSerializer
            return AthleteDetailSerializer

        return UserBaseSerializer


# --------------------------------------------------------------------
#                       ATHLETE INFO
# --------------------------------------------------------------------
class AthleteInfoView(APIView):
    """API для управления персональной информацией атлета."""

    def get_object(self, user_id):
        try:
            user = User.objects.get(id=user_id)
        except User.DoesNotExist:
            return None, Response({"error": "User not found"}, status=404)

        athlete_info, _ = AthleteInfo.objects.get_or_create(user=user)
        return athlete_info, None

    def get(self, request, user_id):
        athlete_info, error = self.get_object(user_id)
        if error:
            return error

        serializer = AthleteInfoSerializer(athlete_info)
        return Response(serializer.data)

    def put(self, request, user_id):
        athlete_info, error = self.get_object(user_id)
        if error:
            return error

        serializer = AthleteInfoSerializer(
            athlete_info,
            data=request.data,
            partial=True,
        )
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=201)

        return Response(serializer.errors, status=400)


class ChallengeViewSet(viewsets.ReadOnlyModelViewSet):
    """API для просмотра выполненных челленджей."""

    queryset = Challenge.objects.all()
    serializer_class = ChallengeSerializer

    def get_queryset(self):
        qs = super().get_queryset()
        athlete_id = self.request.query_params.get("athlete")
        if athlete_id:
            qs = qs.filter(athlete_id=athlete_id)
        return qs


class PositionViewSet(viewsets.ModelViewSet):
    """API для работы с позициями атлетов."""

    queryset = Position.objects.all()
    serializer_class = PositionSerializer

    def get_queryset(self):
        qs = super().get_queryset()
        run_id = self.request.query_params.get("run")
        if run_id:
            qs = qs.filter(run_id=run_id)
        return qs

    def perform_create(self, serializer):
        """
        Сохраняет позицию, рассчитывает скорость и дистанцию,
        а также выполняет сбор ближайших предметов.
        """

        run = serializer.validated_data["run"]

        # 1. Забег должен быть в статусе in_progress
        if run.status != Run.Status.IN_PROGRESS:
            raise serializers.ValidationError(
                "Run must be in progress to record positions"
            )

        # 2. Сохраняем позицию
        position = serializer.save()

        prev = (
            Position.objects.filter(run=run)
            .exclude(pk=position.pk)
            .order_by("-date_time")
            .first()
        )

        if prev is None:
            # первая точка
            position.speed = 0.0
            position.distance = 0.0
        else:
            prev_point = (float(prev.latitude), float(prev.longitude))
            curr_point = (float(position.latitude), float(position.longitude))

            segment_m = geodesic(prev_point, curr_point).meters

            if position.date_time and prev.date_time:
                delta = (position.date_time - prev.date_time).total_seconds()
            else:
                delta = 0

            speed = segment_m / delta if delta > 0 else 0.0
            distance_km = (prev.distance or 0.0) + (segment_m / 1000.0)

            position.speed = round(speed, 2)
            position.distance = round(distance_km, 2)

        position.save(update_fields=["speed", "distance"])

        # 3. Сбор предметов (Collectible Items)
        user = run.athlete
        for item in CollectibleItem.objects.all():
            dist = geodesic(
                (position.latitude, position.longitude),
                (item.latitude, item.longitude),
            ).meters
            if dist <= 100:
                item.collected_by.add(user)


class CollectibleItemView(generics.ListAPIView):
    """API для получения списка коллекционных предметов."""

    queryset = CollectibleItem.objects.all()
    serializer_class = CollectibleItemSerializer


class UploadCollectibleFile(APIView):
    """Загрузка и обработка Excel-файла с коллекционными предметами."""

    def post(self, request):
        file = request.FILES.get("file")
        if not file:
            return Response({"error": "Файл не передан"}, status=400)

        try:
            wb = load_workbook(file)
            sheet = wb.active
        except Exception:
            return Response({"error": "Неверный формат файла"}, status=400)

        invalid_rows = []

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if all(v is None for v in row):
                continue

            # Name, UID, Value, Latitude, Longitude, URL
            name, uid, value, lat, lon, picture = row

            serializer = CollectibleItemSerializer(
                data={
                    "name": name,
                    "uid": uid,
                    "value": value,
                    "latitude": lat,
                    "longitude": lon,
                    "picture": picture,
                }
            )

            if serializer.is_valid():
                serializer.save()
            else:
                invalid_rows.append(list(row))

        return Response(invalid_rows, status=200)
